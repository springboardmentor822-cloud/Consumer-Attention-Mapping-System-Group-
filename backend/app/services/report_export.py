"""
Store performance report export - PDF and Excel.

Scope decision (see chat): the spec's Daily/Weekly/Monthly/Custom Report
framing implies date-range filtering that doesn't actually exist in this
backend - dwell time and shelf scores are always "most recent tracking
run only", with no query-by-date-range anywhere except attractiveness
history (which the scheduler retains for RETENTION_DAYS, unrelated to
this export). Building fake Daily/Weekly/Monthly buttons that all just
export the same current snapshot would misrepresent them as different
data when there's no data behind the distinction. So this is
deliberately ONE export - "current snapshot" - not four.

Every non-real metric already flagged in the UI (mock_metrics on
attractiveness, distinct_visitors as a cross-camera estimate) is
flagged again here, in the exported document itself. A report leaving
the app is exactly the wrong place to quietly drop those caveats -
someone reading a PDF has no access to the dashboard's tooltip text.
"""
import io
import uuid
from datetime import datetime, timezone

from sqlmodel import Session, select
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.core.db import engine
from app.models.camera import Camera
from app.services.compute_dwell_time import compute_dwell_time_data, DwellTimeUnavailable
from app.services.attractiveness_score import compute_attractiveness_scores, AttractivenessScoringUnavailable
from app.services.traffic_analytics_service import compute_zone_traffic


def build_report_data(store_id: uuid.UUID) -> dict:
    with Session(engine) as session:
        cameras = session.exec(select(Camera).where(Camera.store_id == store_id)).all()

    zone_traffic = compute_zone_traffic(store_id)

    dwell_rows: list[dict] = []
    shelf_rows: list[dict] = []
    for cam in cameras:
        try:
            for d in compute_dwell_time_data(cam.id):
                dwell_rows.append({**d, "camera_name": cam.name})
        except DwellTimeUnavailable:
            pass  # no ShelfCameraView / no events for this camera - not an error, just nothing to add

        try:
            # persist=False - this is a read for a report, not a new
            # scoring run. The scheduler (recommendation_scheduler.py)
            # already owns real scoring on its own 15-min cadence;
            # downloading a report shouldn't itself add extra rows to
            # product_attractiveness_scores.
            for s in compute_attractiveness_scores(cam.id, persist=False):
                shelf_rows.append({**s, "camera_name": cam.name})
        except AttractivenessScoringUnavailable:
            pass

    return {
        "generated_at": datetime.now(timezone.utc),
        "cameras_total": len(cameras),
        "cameras_online": sum(1 for c in cameras if c.is_active),
        "zone_traffic": zone_traffic,
        "dwell_rows": dwell_rows,
        "shelf_rows": shelf_rows,
    }


def _pdf_table(data: list[list[str]]) -> Table:
    t = Table(data, hAlign="LEFT")
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


def render_pdf(store, data: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title=f"{store.name} Report")
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"{store.name} \u2014 Store Performance Report", styles["Title"]))
    elements.append(Paragraph(
        f"Generated {data['generated_at'].strftime('%Y-%m-%d %H:%M UTC')} &middot; "
        f"Cameras online: {data['cameras_online']}/{data['cameras_total']} &middot; "
        f"Current snapshot only \u2014 not a date-ranged report (see note below)",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Zone Traffic", styles["Heading2"]))
    zt = [["Zone", "Tracked Events", "Distinct Visitors (est.)"]]
    for z in data["zone_traffic"]:
        zt.append([z["zone_name"], str(z["event_count"]), str(z["distinct_visitors"])])
    elements.append(_pdf_table(zt) if len(zt) > 1 else Paragraph("No zone data yet.", styles["Normal"]))
    elements.append(Paragraph(
        "Distinct Visitors is a conservative max-across-cameras estimate, not an exact headcount \u2014 "
        "there is no cross-camera person re-identification in this system.",
        styles["Italic"],
    ))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Shelf Performance", styles["Heading2"]))
    sp = [["Shelf", "Camera", "Score", "Attention", "Interaction", "Pickup", "Purchase", "Repeat"]]
    for s in data["shelf_rows"]:
        sp.append([
            s["shelf_name"], s["camera_name"], f"{s['final_score']:.3f}",
            f"{s['attention_score']:.2f}", f"{s['interaction_score']:.2f}",
            f"{s['pickup_score']:.2f}", f"{s['purchase_score']:.2f}", f"{s['repeat_score']:.2f}",
        ])
    elements.append(_pdf_table(sp) if len(sp) > 1 else Paragraph("No shelf data yet.", styles["Normal"]))
    elements.append(Paragraph(
        "Only Attention is a real, dwell-time-derived signal. Interaction, Pickup, Purchase, and Repeat "
        "are placeholder values pending real product-interaction tracking \u2014 do not present these "
        "columns as measured shopper behavior.",
        styles["Italic"],
    ))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Dwell Time by Shelf", styles["Heading2"]))
    dt = [["Shelf", "Camera", "Total Seconds", "Distinct Visitors"]]
    for d in data["dwell_rows"]:
        dt.append([d["shelf_name"], d["camera_name"], f"{d['total_seconds']:.1f}", str(d["distinct_visitors"])])
    elements.append(_pdf_table(dt) if len(dt) > 1 else Paragraph("No dwell time data yet.", styles["Normal"]))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph(
        "This report reflects the most recent tracking run per camera at the time of export, not a "
        "specific day/week/month \u2014 the system does not yet support querying dwell time or shelf "
        "scores by date range.",
        styles["Italic"],
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def render_excel(store, data: dict) -> io.BytesIO:
    wb = Workbook()

    ws = wb.active
    ws.title = "Overview"
    ws.append(["Store", store.name])
    ws.append(["Generated", data["generated_at"].strftime("%Y-%m-%d %H:%M UTC")])
    ws.append(["Cameras Online", f"{data['cameras_online']}/{data['cameras_total']}"])
    ws.append(["Scope", "Current snapshot only - not date-ranged (see Store Performance chat/spec notes)"])

    ws2 = wb.create_sheet("Zone Traffic")
    ws2.append(["Zone", "Tracked Events", "Distinct Visitors (est.)"])
    for z in data["zone_traffic"]:
        ws2.append([z["zone_name"], z["event_count"], z["distinct_visitors"]])
    ws2.append([])
    ws2.append(["Note: Distinct Visitors is a conservative max-across-cameras estimate, not an exact headcount."])

    ws3 = wb.create_sheet("Shelf Performance")
    ws3.append(["Shelf", "Camera", "Score", "Attention (real)", "Interaction (mock)",
                "Pickup (mock)", "Purchase (mock)", "Repeat (mock)"])
    for s in data["shelf_rows"]:
        ws3.append([
            s["shelf_name"], s["camera_name"], s["final_score"],
            s["attention_score"], s["interaction_score"], s["pickup_score"],
            s["purchase_score"], s["repeat_score"],
        ])
    ws3.append([])
    ws3.append(["Note: only Attention is real (dwell-time derived). The rest are placeholders."])

    ws4 = wb.create_sheet("Dwell Time")
    ws4.append(["Shelf", "Camera", "Total Seconds", "Distinct Visitors"])
    for d in data["dwell_rows"]:
        ws4.append([d["shelf_name"], d["camera_name"], d["total_seconds"], d["distinct_visitors"]])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
