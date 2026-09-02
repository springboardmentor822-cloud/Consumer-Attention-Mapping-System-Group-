import hashlib
import io
import uuid
from datetime import datetime, time, UTC

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from sqlmodel import Session, select

from app.core.db import engine
from app.core.deps import require_roles
from app.models.campaign import Campaign, CampaignStatus
from app.models.product_attractiveness_score import ProductAttractivenessScore
from app.models.store import Shelf


router = APIRouter()


def _date_start(value):
    return datetime.combine(value, time.min).replace(tzinfo=UTC)


def _date_end(value):
    return datetime.combine(value, time.max).replace(tzinfo=UTC)


def _load_campaign(campaign_id: uuid.UUID):
    with Session(engine) as session:
        campaign = session.get(Campaign, campaign_id)
        if not campaign:
            raise HTTPException(status_code=404, detail="Campaign not found.")

        shelf = session.get(Shelf, campaign.shelf_id)
        if not shelf or shelf.store_id != campaign.store_id:
            raise HTTPException(status_code=409, detail="Campaign shelf/store relationship is invalid.")

        rows = session.exec(
            select(ProductAttractivenessScore)
            .where(
                ProductAttractivenessScore.store_id == campaign.store_id,
                ProductAttractivenessScore.shelf_id == campaign.shelf_id,
                ProductAttractivenessScore.computed_at >= _date_start(campaign.start_date),
                ProductAttractivenessScore.computed_at <= _date_end(campaign.end_date),
            )
            .order_by(ProductAttractivenessScore.computed_at.asc())
        ).all()

    return campaign, shelf, rows


def _campaign_funnel(campaign_id: uuid.UUID) -> dict:
    """
    Marketing funnel (Impressions -> Viewed -> Engaged -> Interested ->
    Converted) has no real backing data anywhere in this system - there is
    no ad-impression or promotional-view tracking, only in-store attention/
    pickup/purchase captured by the camera pipeline. This follows the same
    pattern as app/services/metrics/mock_providers.py: deterministic per
    campaign (hashed, not re-randomized on every call) so the chart doesn't
    visibly jitter between refreshes, and explicitly disclosed via
    is_mock=True rather than presented as observed data.
    """
    digest = hashlib.md5(f"{campaign_id}-funnel".encode()).hexdigest()
    seed = int(digest, 16)

    impressions = 8000 + (seed % 4000)
    stages = [("Impressions", impressions)]
    remaining = impressions
    # Each stage keeps a deterministic 45-75% of the previous stage, so the
    # funnel always narrows (a real funnel never widens) while still
    # varying by campaign.
    for i, label in enumerate(("Viewed", "Engaged", "Interested", "Converted")):
        digit = int(digest[i * 2 : i * 2 + 2], 16)
        keep_ratio = 0.45 + (digit / 255) * 0.30
        remaining = round(remaining * keep_ratio)
        stages.append((label, remaining))

    return {
        "stages": [{"stage": label, "count": count} for label, count in stages],
        "is_mock": True,
        "disclosure": (
            "Impressions/Viewed/Engaged/Interested have no tracking source in this "
            "system (no ad-impression or promotional-view pipeline exists) and are "
            "a deterministic illustrative mock, not observed data. Only in-store "
            "attention, pickup, and purchase (shown elsewhere on this dashboard) are "
            "measured."
        ),
    }


def _analytics_payload(campaign, shelf, rows):
    base = {
        "id": str(campaign.id),
        "name": campaign.name,
        "store_id": str(campaign.store_id),
        "shelf_id": str(campaign.shelf_id),
        "shelf_name": shelf.shelf_name,
        "start_date": campaign.start_date.isoformat(),
        "end_date": campaign.end_date.isoformat(),
        "status": campaign.status.value,
    }

    if not rows:
        return {
            "campaign": base,
            "has_data": False,
            "message": "No attractiveness scores were recorded for this shelf during the campaign date range.",
            "summary": None,
            "trend": [],
            "mock_metrics": [],
            "data_quality": None,
            "promotion_effectiveness": None,
            "engagement": None,
            "conversion": None,
            "funnel": _campaign_funnel(campaign.id),
        }

    latest = rows[-1]
    metrics = ("interaction", "pickup", "purchase", "repeat")
    mock_metrics = sorted({
        metric.strip()
        for row in rows
        for metric in row.mock_metrics.split(",")
        if metric.strip()
    })

    def avg(field):
        values = [getattr(row, field) for row in rows]
        return round(sum(values) / len(values), 3)

    midpoint = len(rows) // 2
    before_rows = rows[:midpoint] if midpoint else rows
    after_rows = rows[midpoint:] if midpoint else rows

    def avg_rows(items, field):
        return round(sum(getattr(row, field) for row in items) / len(items), 3)

    before_final = avg_rows(before_rows, "final_score")
    after_final = avg_rows(after_rows, "final_score")
    before_attention = avg_rows(before_rows, "attention_score")
    after_attention = avg_rows(after_rows, "attention_score")

    before_purchase = avg_rows(before_rows, "purchase_score")
    after_purchase = avg_rows(after_rows, "purchase_score")

    return {
        "campaign": base,
        "has_data": True,
        "summary": {
            "latest_final_score": latest.final_score,
            "average_final_score": avg("final_score"),
            "latest_attention_score": latest.attention_score,
            "average_attention_score": avg("attention_score"),
            "before_final_score": before_final,
            "after_final_score": after_final,
            "final_score_change": round(after_final - before_final, 3),
            "before_attention_score": before_attention,
            "after_attention_score": after_attention,
            "attention_score_change": round(after_attention - before_attention, 3),
            "average_interaction_score": avg("interaction_score"),
            "average_pickup_score": avg("pickup_score"),
            "average_purchase_score": avg("purchase_score"),
            "average_repeat_score": avg("repeat_score"),
            "before_purchase_score": before_purchase,
            "after_purchase_score": after_purchase,
            "purchase_score_change": round(after_purchase - before_purchase, 3),
            "sample_count": len(rows),
        },
        "trend": [
            {
                "computed_at": row.computed_at.isoformat(),
                "final_score": row.final_score,
                "attention_score": row.attention_score,
                "interaction_score": row.interaction_score,
                "pickup_score": row.pickup_score,
                "purchase_score": row.purchase_score,
                "repeat_score": row.repeat_score,
            }
            for row in rows
        ],
        "mock_metrics": mock_metrics,
        "data_quality": {
            "attention_score": "real",
            **{
                f"{metric}_score": "mock" if metric in mock_metrics else "real"
                for metric in metrics
            },
            "final_score": "partially_mocked" if mock_metrics else "real",
        },
        "promotion_effectiveness": {
            "attractiveness_lift": round(after_final - before_final, 3),
            "attention_lift": round(after_attention - before_attention, 3),
            "purchase_proxy_lift": round(after_purchase - before_purchase, 3),
            "method": "first-half vs second-half averages within campaign window",
            "purchase_proxy_is_real": "purchase" not in mock_metrics,
        },
        "engagement": {
            "attention": avg("attention_score"),
            "interaction": avg("interaction_score"),
            "pickup": avg("pickup_score"),
            "repeat": avg("repeat_score"),
            "all_components_observed": not any(m in mock_metrics for m in ("interaction", "pickup", "repeat")),
        },
        "conversion": {
            "attention_to_purchase_proxy": round(
                avg("purchase_score") / avg("attention_score"), 3
            ) if avg("attention_score") else None,
            "purchase_score": avg("purchase_score"),
            "is_observed_conversion": "purchase" not in mock_metrics,
        },
        "funnel": _campaign_funnel(campaign.id),
    }


@router.get("/{campaign_id}/analytics")
def campaign_analytics(
    campaign_id: uuid.UUID,
    current_user=Depends(require_roles("MarketingManager", "SuperAdmin")),
):
    campaign, shelf, rows = _load_campaign(campaign_id)
    return _analytics_payload(campaign, shelf, rows)


@router.get("/{campaign_id}/report")
def campaign_report(
    campaign_id: uuid.UUID,
    current_user=Depends(require_roles("MarketingManager", "SuperAdmin")),
):
    campaign, shelf, rows = _load_campaign(campaign_id)
    analytics = _analytics_payload(campaign, shelf, rows)

    recommendations = []
    if analytics["has_data"]:
        s = analytics["summary"]
        if s["average_attention_score"] - s["average_pickup_score"] > 0.15:
            recommendations.append(
                "Attention is materially higher than pickup; review placement, shelf messaging, and product interaction friction."
            )
        if s["average_final_score"] >= 0.70:
            recommendations.append(
                "This shelf is a strong attractiveness benchmark; preserve its placement and campaign treatment."
            )
        else:
            recommendations.append(
                "Attractiveness is below the strong-performance threshold; review placement and attention drivers."
            )
        if analytics["mock_metrics"]:
            recommendations.append(
                "Do not use mocked interaction/pickup/purchase/repeat metrics for business claims until real providers replace them."
            )

    return {
        **analytics,
        "recommendations": recommendations,
        "report_scope": "campaign date range and the selected shelf",
        "limitations": [
            "There is no campaign-linked sales/revenue source.",
            "Traffic analytics are latest tracking-run analytics, not date-ranged campaign attribution.",
            "Purchase, pickup, interaction and repeat may be mocked.",
            "Cross-camera re-identification is not available.",
        ],
    }


def _safe_filename(name: str):
    cleaned = "".join(c for c in name.strip() if c.isalnum() or c in (" ", "-", "_")).strip()
    return cleaned or "campaign"


@router.get("/{campaign_id}/export")
def export_campaign_report(
    campaign_id: uuid.UUID,
    format: str = "pdf",
    current_user=Depends(require_roles("MarketingManager", "SuperAdmin")),
):
    if format not in ("pdf", "excel"):
        raise HTTPException(status_code=400, detail="format must be 'pdf' or 'excel'.")

    campaign, shelf, rows = _load_campaign(campaign_id)
    report = _analytics_payload(campaign, shelf, rows)

    # _analytics_payload() is intentionally shared by the JSON analytics
    # endpoint and does not include report-only metadata. Exports, however,
    # include a Limitations sheet/section, so add the same disclosures here
    # before rendering either format.
    report["limitations"] = [
        "There is no campaign-linked sales/revenue source.",
        "Traffic analytics are latest tracking-run analytics, not date-ranged campaign attribution.",
        "Purchase, pickup, interaction and repeat may be mocked.",
        "Cross-camera re-identification is not available.",
    ]

    filename_base = _safe_filename(campaign.name)

    if format == "excel":
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Campaign Summary"
        ws.append(["Campaign", campaign.name])
        ws.append(["Shelf", shelf.shelf_name])
        ws.append(["Start", campaign.start_date.isoformat()])
        ws.append(["End", campaign.end_date.isoformat()])
        ws.append(["Status", campaign.status.value])
        ws.append([])
        ws.append(["Metric", "Value", "Quality"])
        if report["summary"]:
            summary = report["summary"]
            ws.append(["Average attractiveness", summary["average_final_score"], report["data_quality"]["final_score"]])
            ws.append(["Average attention", summary["average_attention_score"], report["data_quality"]["attention_score"]])
            ws.append(["Average interaction", summary["average_interaction_score"], report["data_quality"]["interaction_score"]])
            ws.append(["Average pickup", summary["average_pickup_score"], report["data_quality"]["pickup_score"]])
            ws.append(["Average purchase", summary["average_purchase_score"], report["data_quality"]["purchase_score"]])
            ws.append(["Average repeat", summary["average_repeat_score"], report["data_quality"]["repeat_score"]])

        ws2 = wb.create_sheet("Trend")
        ws2.append(["Computed At", "Final", "Attention", "Interaction", "Pickup", "Purchase", "Repeat"])
        for row in report["trend"]:
            ws2.append([
                row["computed_at"], row["final_score"], row["attention_score"],
                row["interaction_score"], row["pickup_score"],
                row["purchase_score"], row["repeat_score"],
            ])

        ws3 = wb.create_sheet("Limitations")
        for item in report["limitations"]:
            ws3.append([item])

        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}_campaign_report.xlsx"'},
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title=f"{campaign.name} Campaign Report")
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"{campaign.name} — Campaign Report", styles["Title"]),
        Paragraph(
            f"Shelf: {shelf.shelf_name} · {campaign.start_date.isoformat()} to {campaign.end_date.isoformat()} · Status: {campaign.status.value}",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]

    if report["summary"]:
        s = report["summary"]
        table = Table([
            ["Metric", "Value", "Quality"],
            ["Average attractiveness", f"{s['average_final_score']:.3f}", report["data_quality"]["final_score"]],
            ["Average attention", f"{s['average_attention_score']:.3f}", report["data_quality"]["attention_score"]],
            ["Average interaction", f"{s['average_interaction_score']:.3f}", report["data_quality"]["interaction_score"]],
            ["Average pickup", f"{s['average_pickup_score']:.3f}", report["data_quality"]["pickup_score"]],
            ["Average purchase", f"{s['average_purchase_score']:.3f}", report["data_quality"]["purchase_score"]],
            ["Average repeat", f"{s['average_repeat_score']:.3f}", report["data_quality"]["repeat_score"]],
            ["Attractiveness lift", f"{s['final_score_change']:+.3f}", "derived"],
            ["Attention lift", f"{s['attention_score_change']:+.3f}", "derived"],
        ])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))
    else:
        elements.append(Paragraph("No attractiveness scores were recorded during the campaign window.", styles["Normal"]))

    elements.append(Paragraph("Recommendations", styles["Heading2"]))
    recommendations = []
    if report["has_data"]:
        s = report["summary"]
        if s["average_attention_score"] - s["average_pickup_score"] > 0.15:
            recommendations.append("Attention is materially higher than pickup; review placement, shelf messaging, and interaction friction.")
        if s["average_final_score"] >= 0.70:
            recommendations.append("Use this shelf as a strong attractiveness benchmark.")
        else:
            recommendations.append("Review placement and attention drivers because attractiveness is below 70%.")
    if report["mock_metrics"]:
        recommendations.append("Mocked metrics must not be presented as observed shopper behavior.")
    for rec in recommendations or ["No data-backed recommendation available."]:
        elements.append(Paragraph(f"• {rec}", styles["Normal"]))
        elements.append(Spacer(1, 4))

    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Limitations", styles["Heading2"]))
    for item in report["limitations"]:
        elements.append(Paragraph(f"• {item}", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}_campaign_report.pdf"'},
    )
