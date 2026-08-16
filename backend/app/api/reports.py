from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from typing import Optional, Any
import datetime
from app.core.database import get_db
from app.api.auth import get_current_user, get_user_email
from app.models import Store, Shelf
from app.utils.logging import get_structured_logger
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.services.analytics_service import AnalyticsService
from app.models.camera_event import CameraEvent
from app.models.notification import Notification
from app.models.product import Product

logger = get_structured_logger("reports_api")
router = APIRouter()

@router.get(
    "/export/pdf",
    summary="Export PDF Report",
    description="Generates and downloads a PDF performance report covering registered store footprint details.",
    response_class=StreamingResponse
)
def export_pdf(
    store_id: Optional[str] = Query(None, description="Optional store filter ID"),
    db: Session = Depends(get_db),
    current_user: Any = Depends(get_current_user)
):
    user_email = get_user_email(current_user)
    logger.info(f"User {user_email} requested PDF export", extra={"store_id": store_id, "user": user_email})
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        story = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#4f46e5'),
            spaceAfter=15
        )

        # Title
        story.append(Paragraph("Consumer Attention Mapping System", title_style))
        story.append(Paragraph("Executive Performance & Attention Intelligence Report", styles['Heading2']))
        story.append(Spacer(1, 15))

        # Store overview
        stores_query = db.query(Store)
        if store_id:
            stores_query = stores_query.filter(Store.id == store_id)
        stores = stores_query.all()

        story.append(Paragraph("Active Store Footprint Details", styles['Heading3']))
        story.append(Spacer(1, 8))

        table_data = [["Store ID", "Name", "Location/Address", "Shelves"]]
        for store in stores:
            shelf_count = db.query(Shelf).filter(Shelf.store_id == store.id).count()
            table_data.append([store.id, store.name, store.address, str(shelf_count)])

        t = Table(table_data, colWidths=[150, 150, 150, 82])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f8fafc')),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f1f5f9')]),
        ]))
        story.append(t)
        story.append(Spacer(1, 15))

        # Add visitor & traffic stats
        if store_id:
            story.append(Paragraph("Visitor Traffic & Dwell Statistics", styles['Heading3']))
            story.append(Spacer(1, 8))
            dwell_data = AnalyticsService.get_dwell_metrics(db, store_id)

            traffic_data = [
                ["Metric", "Value"],
                ["Total Customer Sessions", str(dwell_data.get("total_sessions", 0))],
                ["Average Dwell Duration", f"{dwell_data.get('average_duration_seconds', 0.0):.1f} seconds"],
                ["Longest Shopping Session", f"{dwell_data.get('longest_session', 0.0):.1f} seconds"],
                ["Shortest Shopping Session", f"{dwell_data.get('shortest_session', 0.0):.1f} seconds"]
            ]
            t_traffic = Table(traffic_data, colWidths=[250, 282])
            t_traffic.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#64748b')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            story.append(t_traffic)
            story.append(Spacer(1, 15))

            # Shelf / Zone performance
            story.append(Paragraph("Aisle Zone Attention & Dwell", styles['Heading3']))
            story.append(Spacer(1, 8))
            zone_data = AnalyticsService.get_zone_metrics(db, store_id)
            zone_table = [["Zone ID", "Zone Name", "Visits", "Average Attention Score"]]
            for zd in zone_data:
                zone_table.append([
                    str(zd.get("zone_id", "")),
                    zd.get("zone_name", ""),
                    str(zd.get("zone_visits", 0)),
                    f"{zd.get('average_attention_score', 0.0):.1f}"
                ])
            t_zones = Table(zone_table, colWidths=[150, 150, 120, 112])
            t_zones.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0ea5e9')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            story.append(t_zones)
            story.append(Spacer(1, 15))

            # Product Engagement Metrics
            story.append(Paragraph("Product Engagement & Conversion Rates", styles['Heading3']))
            story.append(Spacer(1, 8))
            product_data = AnalyticsService.get_product_metrics(db, store_id)
            product_table = [["Product Name", "Views", "Pickups", "Compares", "Purchases", "Conversion Rate"]]
            for pd in product_data:
                product_table.append([
                    pd.get("product_name", ""),
                    str(pd.get("views", 0)),
                    str(pd.get("pickups", 0)),
                    str(pd.get("compares", 0)),
                    str(pd.get("purchases", 0)),
                    f"{pd.get('conversion_rate', 0.0):.1f}%"
                ])
            t_prod = Table(product_table, colWidths=[150, 70, 70, 70, 70, 102])
            t_prod.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            story.append(t_prod)
            story.append(Spacer(1, 15))

            # Active/Recent Alerts
            story.append(Paragraph("Active Surveillance System Alerts", styles['Heading3']))
            story.append(Spacer(1, 8))
            from app.models.camera import Camera
            cameras = db.query(Camera).filter(Camera.store_id == store_id).all()
            cam_ids = [c.id for c in cameras]
            alerts = []
            if cam_ids:
                alerts = db.query(CameraEvent).filter(CameraEvent.camera_id.in_(cam_ids)).order_by(CameraEvent.timestamp.desc()).limit(10).all()

            alert_table = [["Timestamp", "Alert Type", "Details"]]
            for al in alerts:
                alert_table.append([
                    al.timestamp.strftime("%Y-%m-%d %H:%M:%S") if al.timestamp else "",
                    al.event_type,
                    al.details or ""
                ])
            if len(alert_table) == 1:
                alert_table.append(["-", "No recent alerts recorded", "-"])

            t_alerts = Table(alert_table, colWidths=[100, 100, 332])
            t_alerts.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f43f5e')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            story.append(t_alerts)

        doc.build(story)
        buffer.seek(0)
        logger.info(f"PDF export compiled successfully for user {user_email}")
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=attention_intelligence_report.pdf"}
        )
    except Exception as e:
        logger.error(f"Failed to export PDF: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error generating report")


@router.get(
    "/export/excel",
    summary="Export Excel Report",
    description="Generates and downloads an Excel performance spreadsheet detailing store layouts and metrics.",
    response_class=StreamingResponse
)
def export_excel(
    store_id: Optional[str] = Query(None, description="Optional store filter ID"),
    db: Session = Depends(get_db),
    current_user: Any = Depends(get_current_user)
):
    user_email = get_user_email(current_user)
    logger.info(f"User {user_email} requested Excel export", extra={"store_id": store_id, "user": user_email})
    try:
        wb = openpyxl.Workbook()

        # Tab 1: Store Footprint
        ws1 = wb.active
        ws1.title = "Store Overview"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="4F46E5")
        cell_font = Font(name="Calibri", size=11)
        fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        align = Alignment(horizontal="center", vertical="center")

        ws1['A1'] = "Consumer Attention Mapping System - Performance Report"
        ws1['A1'].font = title_font
        ws1.merge_cells('A1:D1')
        ws1.row_dimensions[1].height = 30

        ws1.append([]) # spacer

        headers1 = ["Store ID", "Name", "Location", "Total Shelves"]
        ws1.append(headers1)
        ws1.row_dimensions[3].height = 20

        for col_idx, h in enumerate(headers1, 1):
            cell = ws1.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = align

        stores_query = db.query(Store)
        if store_id:
            stores_query = stores_query.filter(Store.id == store_id)
        stores = stores_query.all()

        for store in stores:
            shelf_count = db.query(Shelf).filter(Shelf.store_id == store.id).count()
            ws1.append([store.id, store.name, store.address, shelf_count])

        for row in range(4, len(stores) + 4):
            for col in range(1, 5):
                ws1.cell(row=row, column=col).font = cell_font

        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Tab 2: Traffic & Dwell (If store selected)
        if store_id:
            ws2 = wb.create_sheet(title="Traffic & Dwell Metrics")
            ws2['A1'] = "Store Traffic & Dwell Statistics"
            ws2['A1'].font = title_font
            ws2.merge_cells('A1:B1')
            ws2.append([])

            headers2 = ["Metric", "Value"]
            ws2.append(headers2)
            for col_idx, h in enumerate(headers2, 1):
                cell = ws2.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
                cell.alignment = align

            dwell_data = AnalyticsService.get_dwell_metrics(db, store_id)
            ws2.append(["Total Customer Sessions", dwell_data.get("total_sessions", 0)])
            ws2.append(["Average Dwell Duration (seconds)", dwell_data.get("average_duration_seconds", 0.0)])
            ws2.append(["Longest Shopping Session (seconds)", dwell_data.get("longest_session", 0.0)])
            ws2.append(["Shortest Shopping Session (seconds)", dwell_data.get("shortest_session", 0.0)])

            for col in ws2.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws2.column_dimensions[col_letter].width = max(max_len + 3, 15)

            # Tab 3: Zone performance
            ws3 = wb.create_sheet(title="Shelf & Zone Performance")
            ws3['A1'] = "Aisle Zone Attention Details"
            ws3['A1'].font = title_font
            ws3.merge_cells('A1:D1')
            ws3.append([])

            headers3 = ["Zone ID", "Zone Name", "Zone Visits", "Average Attention Score"]
            ws3.append(headers3)
            for col_idx, h in enumerate(headers3, 1):
                cell = ws3.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
                cell.alignment = align

            zone_data = AnalyticsService.get_zone_metrics(db, store_id)
            for zd in zone_data:
                ws3.append([
                    zd.get("zone_id", ""),
                    zd.get("zone_name", ""),
                    zd.get("zone_visits", 0),
                    zd.get("average_attention_score", 0.0)
                ])

            for col in ws3.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws3.column_dimensions[col_letter].width = max(max_len + 3, 15)

            # Tab 4: Product engagement
            ws4 = wb.create_sheet(title="Product Engagement")
            ws4['A1'] = "Product Interaction and Conversion Rates"
            ws4['A1'].font = title_font
            ws4.merge_cells('A1:F1')
            ws4.append([])

            headers4 = ["Product Name", "Views", "Pickups", "Compares", "Purchases", "Conversion Rate"]
            ws4.append(headers4)
            for col_idx, h in enumerate(headers4, 1):
                cell = ws4.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.alignment = align

            product_data = AnalyticsService.get_product_metrics(db, store_id)
            for pd in product_data:
                ws4.append([
                    pd.get("product_name", ""),
                    pd.get("views", 0),
                    pd.get("pickups", 0),
                    pd.get("compares", 0),
                    pd.get("purchases", 0),
                    pd.get("conversion_rate", 0.0)
                ])

            for col in ws4.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws4.column_dimensions[col_letter].width = max(max_len + 3, 15)

            # Tab 5: Surveillance alerts
            ws5 = wb.create_sheet(title="System Alerts")
            ws5['A1'] = "Active Security & Hardware Alerts Log"
            ws5['A1'].font = title_font
            ws5.merge_cells('A1:C1')
            ws5.append([])

            headers5 = ["Timestamp", "Alert Type", "Details"]
            ws5.append(headers5)
            for col_idx, h in enumerate(headers5, 1):
                cell = ws5.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = PatternFill(start_color="F43F5E", end_color="F43F5E", fill_type="solid")
                cell.alignment = align

            from app.models.camera import Camera
            cameras = db.query(Camera).filter(Camera.store_id == store_id).all()
            cam_ids = [c.id for c in cameras]
            alerts = []
            if cam_ids:
                alerts = db.query(CameraEvent).filter(CameraEvent.camera_id.in_(cam_ids)).order_by(CameraEvent.timestamp.desc()).limit(20).all()

            for al in alerts:
                ws5.append([
                    al.timestamp.strftime("%Y-%m-%d %H:%M:%S") if al.timestamp else "",
                    al.event_type,
                    al.details or ""
                ])

            for col in ws5.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws5.column_dimensions[col_letter].width = max(max_len + 3, 15)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(f"Excel export compiled successfully for user {user_email}")
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=attention_intelligence_report.xlsx"}
        )
    except Exception as e:
        logger.error(f"Failed to export Excel: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error generating report")
